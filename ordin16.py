import os
import xlsxwriter
from openpyxl import load_workbook
from tkinter import *
from tkinter import ttk, filedialog, messagebox
from pprint import pprint
import json

# Get current directory.
cwd = os.getcwd()
# Create a filepath variable.
filepath = ""
# Create a file_save_to variable.
file_save_to = ""
# Create a month variable.
month = ""
# Get database path.
database = os.path.join(cwd, "database\\all_suppliers.json")
with open(database) as file:
    data = json.load(file)
# Define a list with all the cells that have to be bordered.
bordered_cells = ["A4", "A5", "A6", "A7", "A8", "A9",
                  "B4", "B5", "B6", "B7", "B8", "B9",
                  "C4", "C5", "C6", "C7", "C8", "C9",
                  "D4", "D5", "D6", "D7", "D8", "D9",
                  "E4", "E5", "E6", "E7", "E8", "E9",
                  "F4", "F5", "F6", "F7", "F8", "F9",
                  "G4", "G5", "G6", "G7", "G8", "G9",
                  "H4", "H5", "H6", "H7", "H8", "H9",
                  "I4", "I5", "I6", "I7", "I8", "I9",
                  "J4", "J5", "J6", "J7", "J8", "J9",
                  "K4", "K5", "K6", "K7", "K8", "K9",
                  "L4", "L5", "L6", "L7", "L8", "L9",
                  "M4", "M5", "M6", "M7", "M8", "M9",
                  "N4", "N5", "N6", "N7", "N8", "N9",
                  "O4", "O5", "O6", "O7", "O8", "O9",
                 ]
# Define a dict with all months and their respective cells.
months_dict = {"Ianuarie":
                   {"Cantitatea Facturata(MWh)": "F7",
                    "Cantitate GMOIS(al fin)": ["F5", "F6"]},
                "Februarie":
                   {"Cantitatea Facturata(MWh)": "G7",
                    "Cantitate GMOIS(al fin)": ["G5", "G6"]},
                "Martie":
                   {"Cantitatea Facturata(MWh)": "H7",
                    "Cantitate GMOIS(al fin)": ["H5", "H6"]},
                "Aprilie":
                   {"Cantitatea Facturata(MWh)": "I7",
                    "Cantitate GMOIS(al fin)": ["I5", "I6"]},
                "Mai":
                   {"Cantitatea Facturata(MWh)": "J7",
                    "Cantitate GMOIS(al fin)": ["J5", "J6"]},
                "Iunie":
                   {"Cantitatea Facturata(MWh)": "K7",
                    "Cantitate GMOIS(al fin)": ["K5", "K6"]},
                "Iulie":
                   {"Cantitatea Facturata(MWh)": "L7",
                    "Cantitate GMOIS(al fin)": ["L5", "L6"]},
                "August":
                   {"Cantitatea Facturata(MWh)": "M7",
                    "Cantitate GMOIS(al fin)": ["M5", "M6"]},
                "Septembrie":
                   {"Cantitatea Facturata(MWh)": "N7",
                    "Cantitate GMOIS(al fin)": ["N5", "N6"]},
                "Octombrie":
                   {"Cantitatea Facturata(MWh)": "C7",
                    "Cantitate GMOIS(al fin)": ["C5", "C6"]},
                "Noiembrie":
                   {"Cantitatea Facturata(MWh)": "D7",
                    "Cantitate GMOIS(al fin)": ["D5", "D6"]},
                "Decembrie":
                   {"Cantitatea Facturata(MWh)": "E7",
                    "Cantitate GMOIS(al fin)": ["E5", "E6"]},
            }


def month_selected(event):
    """
    A function that takes the information from a combobox.
    :param event: get()
    :return: chosen month
    """
    global month
    month = combo.get()
    return month

def open_file():
    """
    A function to return a path
    :return: file path.
    """
    global filepath
    filepath = filedialog.askopenfile(initialdir=cwd, title="Select a file...", filetypes=[("New Excel", "*.xlsx")])


def save_to():
    """
    A function which asks the user where to save the files.
    :return: the path where to save the file(s).
    """
    global file_save_to
    file_save_to = filedialog.askdirectory(initialdir=cwd, title="Save to...")

def generate_files():
    """
    A function that generates all files according to supplier
    Creates a new excel file with a sheetname as the supplier NU(Newtork User) code containing all the data required
    for monthly reporting.
    """
    # If the user has chosen a month we can go further.
    if month:
        # Then the user needs to choose the import filepath.
        if filepath:
            path = str(filepath).split("'")[1]
            wb = load_workbook(path)
            wb.active = wb["Sheet1"]
            ws = wb.active
            suppliers = {}
            for i in range(3, ws.max_row):
                cell = ws.cell(row=i, column=2)
                try:
                    if cell.value == cell.value.upper() and len(cell.value) == 6:
                        suppliers[cell.value] = {"Cantitatea Facturata(MWh)": ws.cell(row=i, column=3).value,
                                                 "Cantitate GMOIS(al fin)": ws.cell(row=i, column=4).value}
                except Exception as e:
                    messagebox.showwarning("Warning!", f"The folowing error has occured: {e}")
            # Make the user select a directory where to save the generated files.
            if file_save_to == "":
                messagebox.showwarning("Warning!",
                                       "Use the 'Save files to...' button to select where to save the files.")
            else:
                # Change directory according to file_save_to.
                os.chdir(file_save_to)
                # Get the number of files in that directory.
                before_dir = len(os.listdir(file_save_to))
                # Loop through the final_suppliers and suppliers to find the
                for key, value in suppliers.items():
                    # Write to a new excel file.
                    wb = xlsxwriter.Workbook(f"{key}-{month}.xlsx")
                    # Add a sheet named by supplier's name
                    ws = wb.add_worksheet(key)
                    # Set the format.
                    bordered_number_format = wb.add_format({"num_format": "#,##0.000000",
                                                            "border": 1})
                    red_bordered_number_format = wb.add_format({"num_format": "#,##0.000000",
                                                            "border": 1,
                                                                "font_color": "red"})
                    bold_format = wb.add_format({"bold": True})
                    bold_and_border = wb.add_format({"bold": True, "border": 1})
                    border_format = wb.add_format({"border": 1})
                    try:
                        for x in bordered_cells:
                            # This block mostly formats the cells, plus a few adjustments.
                            ws.write(x, "", border_format)
                            ws.write("A2", "OD: SC GAZ VEST SA", bold_format)
                            ws.write("A3", f"UR: {data[key]}", bold_format)
                            if x == "A4":
                                ws.write(x, "", bold_and_border)
                            if x == "B4":
                                ws.write(x, "UR", bold_and_border)
                            if x == "C4":
                                ws.write(x, "Oct-21", bold_and_border)
                            if x == "D4":
                                ws.write(x, "Nov-21", bold_and_border)
                            if x == "E4":
                                ws.write(x, "Dec-21", bold_and_border)
                            if x == "F4":
                                ws.write(x, "Ian-22", bold_and_border)
                            if x == "G4":
                                ws.write(x, "Feb-22", bold_and_border)
                            if x == "H4":
                                ws.write(x, "Mar-22", bold_and_border)
                            if x == "I4":
                                ws.write(x, "Apr-22", bold_and_border)
                            if x == "J4":
                                ws.write(x, "May-22", bold_and_border)
                            if x == "K4":
                                ws.write(x, "Jun-22", bold_and_border)
                            if x == "L4":
                                ws.write(x, "Jul-22", bold_and_border)
                            if x == "M4":
                                ws.write(x, "Aug-22", bold_and_border)
                            if x == "N4":
                                ws.write(x, "Sep-22", bold_and_border)
                            if x == "O4":
                                ws.write(x, "Total", bold_and_border)
                            if x == "A5":
                                ws.write(x, "Alocari finale", border_format)
                            if x == "A6":
                                ws.write(x, "Suma alocarilor zilnice", border_format)
                            if x == "A7":
                                ws.write(x, "Cantitate distribuita", border_format)
                            if x == "A8":
                                ws.write(x, "Alocari finale-Cantitate distribuita", border_format)
                            if x == "A9":
                                ws.write(x, "Alocari Finale-Alocari zilnice", border_format)
                            if x == "B5":
                                ws.write(x, key, border_format)
                            if x == "B6":
                                ws.write(x, key, border_format)
                            if x == "B7":
                                ws.write(x, key, border_format)
                            if x == "C8":
                                ws.write(x, "=C5-C7", red_bordered_number_format)
                            if x == "C9":
                                ws.write(x, "=C5-C6", red_bordered_number_format)
                            if x == "D8":
                                ws.write(x, "=D5-D7", red_bordered_number_format)
                            if x == "D9":
                                ws.write(x, "=D5-D6", red_bordered_number_format)
                            if x == "E8":
                                ws.write(x, "=E5-E7", red_bordered_number_format)
                            if x == "E9":
                                ws.write(x, "=E5-E6", red_bordered_number_format)
                            if x == "F8":
                                ws.write(x, "=F5-F7", red_bordered_number_format)
                            if x == "F9":
                                ws.write(x, "=F5-F6", red_bordered_number_format)
                            if x == "G8":
                                ws.write(x, "=G5-G7", red_bordered_number_format)
                            if x == "G9":
                                ws.write(x, "=G5-G6", red_bordered_number_format)
                            if x == "H8":
                                ws.write(x, "=H5-H7", red_bordered_number_format)
                            if x == "H9":
                                ws.write(x, "=H5-H6", red_bordered_number_format)
                            if x == "I8":
                                ws.write(x, "=I5-I7", red_bordered_number_format)
                            if x == "I9":
                                ws.write(x, "=I5-I6", red_bordered_number_format)
                            if x == "J8":
                                ws.write(x, "=J5-J7", red_bordered_number_format)
                            if x == "J9":
                                ws.write(x, "=J5-J6", red_bordered_number_format)
                            if x == "K8":
                                ws.write(x, "=K5-K7", red_bordered_number_format)
                            if x == "K9":
                                ws.write(x, "=K5-K6", red_bordered_number_format)
                            if x == "L8":
                                ws.write(x, "=L5-L7", red_bordered_number_format)
                            if x == "L9":
                                ws.write(x, "=L5-L6", red_bordered_number_format)
                            if x == "M8":
                                ws.write(x, "=M5-M7", red_bordered_number_format)
                            if x =="M9":
                                ws.write(x, "=M5-M6", red_bordered_number_format)
                            if x == "N8":
                                ws.write(x, "=N5-N7", red_bordered_number_format)
                            if x == "N9":
                                ws.write(x, "=N5-N6", red_bordered_number_format)
                            if x == "O5":
                                ws.write(x, "=SUM(C5:N5)", bordered_number_format)
                            if x == "O6":
                                ws.write(x, "=SUM(C6:N6)", bordered_number_format)
                            if x == "O7":
                                ws.write(x, "=SUM(C7:N7)", bordered_number_format)
                            if x == "O8":
                                ws.write(x, "=SUM(C8:N8)", red_bordered_number_format)
                            if x == "O9":
                                ws.write(x, "=SUM(C9:N9)", red_bordered_number_format)
                        # Set where to write the data from final_suppliers.
                        ws.write(str(months_dict[month]["Cantitatea Facturata(MWh)"]),
                                 value["Cantitatea Facturata(MWh)"], bordered_number_format)
                        cant_fact_cell = months_dict[month]["Cantitatea Facturata(MWh)"]
                        cant_al_fin = months_dict[month]["Cantitate GMOIS(al fin)"][0]
                        cant_al_fin1 = months_dict[month]["Cantitate GMOIS(al fin)"][1]
                        ws.write(cant_fact_cell, value["Cantitatea Facturata(MWh)"], bordered_number_format)
                        ws.write(cant_al_fin, value["Cantitate GMOIS(al fin)"], bordered_number_format)
                        ws.write(cant_al_fin1, value["Cantitate GMOIS(al fin)"], bordered_number_format)
                        wb.close()
                    except KeyError as e:
                        messagebox.showwarning("Warning!",
                                               f"We encountered a problem when creating file for supplier: {e}")
                    except Exception as e:
                        messagebox.showwarning("Warning!", f"The following exception has occured: {e}")
                # Get the number of files created after generating.
                after_dir = len(os.listdir(file_save_to))
                if after_dir - before_dir == 0:
                    messagebox.showinfo("Information!", "No files were created.")
                # Show a message that informs the user how many files have been created.
                else:
                    messagebox.showinfo("Information!", f"Succesfully created {after_dir - before_dir} files.")
        # If no filepath to import a file from has been give, return a message.
        else:
            messagebox.showwarning("Warning!", "You have to upload a file to convert.")
    # If there hasn't been selected a month, return a message.
    else:
        messagebox.showwarning("Warning!", "You have to select the month.")
    pprint(data)


def new_supplier():
    """
    A function that gives the user the ability to enter a new supplier into the database in case it doesn't exist.
    It opens a new window with needed entry boxes for adding the new supplier.
    """

    global key_entry
    global value_entry
    global new_root

    # Create a new window.
    new_root = Tk()
    # Set the minimum size
    new_root.minsize(width=400, height=300)
    # Set the actual size.
    new_root.geometry("400x300")
    # Set the title.
    new_root.title("Add new supplier...")
    # Set an icon
    new_root.iconbitmap('images\\GazVest.ico')
    # Set up the grid.
    new_root.columnconfigure(0, weight=1)
    # Key label and entry.
    nu_label_descr = Label(new_root, text="Insert a new Network User.")
    nu_label_descr.grid(row=0, column=0, sticky=W)
    key_label_descr = Label(new_root, text="Key=Network User Code(ex. EGROMS).")
    key_label_descr.grid(row=1, column=0, sticky=W)
    value_label_descr = Label(new_root, text="Value=Supplier name(ex. E.ON ENERGIE ROMANIA.)")
    value_label_descr.grid(row=2, column=0, sticky=W)
    empty_row = Label(new_root, text="")
    empty_row.grid(row=3, column=0, pady=20)
    key_entry_label = Label(new_root, text="Insert the KEY:")
    key_entry_label.grid(row=4, column=0, sticky=W)
    key_entry = Entry(new_root)
    key_entry.grid(row=4, column=0, )
    value_entry_label = Label(new_root, text="Insert the VALUE:")
    value_entry_label.grid(row=5, column=0, sticky=W)
    value_entry = Entry(new_root)
    value_entry.grid(row=5, column=0, pady=5)
    add_new_supl_button = Button(new_root, text="Add to database", command=add_new_supplier)
    add_new_supl_button.grid(row=6, column=0, pady=10)



def add_new_supplier():
    """
    A function that adds new supplier to database(json file).
    If the Network User COde already exists it returns an error.
    """
    # Check if the entered key is already in the database. If it is, show a message and close the window.
    if key_entry.get().upper() in data.keys():
        messagebox.showwarning("Warning!", f"Network User Code: {key_entry.get().upper()} already exists. Cannot be added.")
        new_root.destroy()
    # If no values are added to key or to value entry box, return an error message and close the window.
    if key_entry.get() == "" or value_entry.get() == "":
        messagebox.showwarning("Warning!",
                               f"You forgot to type a key or a value. You have to fill in both fields.")
        new_root.destroy()
    #
    else:
        with open(database, 'w') as file:
            data[key_entry.get().upper()] = value_entry.get().upper()
            json.dump(data, file)
            messagebox.showinfo("Information!",
                                   f"Network User Code: {key_entry.get().upper()} with supplier name: {value_entry.get().upper()}"
                                   f" has been added to the database successfully!")
        new_root.destroy()



def delete_supplier():
    """
    A function that creates a new window where user can add the data needed to delete a supplier.

    """
    global key_entry_delete
    global new_root_delete

    # Create a new window.
    new_root_delete = Tk()
    # Set the minimum size
    new_root_delete.minsize(width=400, height=300)
    # Set the actual size.
    new_root_delete.geometry("400x300")
    # Set the title.
    new_root_delete.title("Delete Supplier...")
    # Set an icon
    new_root_delete.iconbitmap('images\\GazVest.ico')
    # Set up the grid.
    new_root_delete.columnconfigure(0, weight=1)
    # Key label and entry.
    nu_label_descr = Label(new_root_delete, text="Delete a Network User.")
    nu_label_descr.grid(row=0, column=0, sticky=W)
    key_label_descr = Label(new_root_delete, text="Key=Network User Code to delete.(ex. EGROMS).")
    key_label_descr.grid(row=1, column=0, sticky=W)
    empty_row = Label(new_root_delete, text="")
    empty_row.grid(row=3, column=0, pady=20)
    key_entry_label = Label(new_root_delete, text="Insert the KEY:")
    key_entry_label.grid(row=4, column=0, sticky=W)
    key_entry_delete = Entry(new_root_delete)
    key_entry_delete.grid(row=4, column=0, )
    delete_button = Button(new_root_delete, text="Delete supplier", command=delete_sup)
    delete_button.grid(row=6, column=0, pady=10)


def delete_sup():
    """
        A function that deletes a supplier from database(json file).
        If the Network User Code doesn't exist it returns an error.
        """
    # If entered key is found in the database, delete it and close the window.
    if key_entry_delete.get().upper() in data.keys():
        with open(database, 'w') as file:
            messagebox.showinfo("Information!",
                                f"Network User Code: {key_entry_delete.get().upper()} with supplier name: {data[key_entry_delete.get().upper()]}"
                                f" has been deleted from the database successfully!")
            del data[key_entry_delete.get().upper()]
            json.dump(data, file)
            new_root_delete.destroy()
    # If eneterd key isn't in our database, return an error and close the window.
    else:
        messagebox.showwarning("Warning!",
                               f"Network User Code: {key_entry_delete.get().upper()} doesn't exist. Cannot be deleted.")
        new_root_delete.destroy()


def show_all_suppliers():
    """
    A function that show all records in the database.
    """
    # Create a new window.
    show_suppliers_window = Tk()
    # Create a treeview color.
    style = ttk.Style()
    style.theme_use("default")
    style.configure("Treeview",
                    background="#D3D3D3",
                    foreground="black",
                    fieldbackground="#D3D3D3")
    style.map("Treeview",
              background=[("selected", "#D347083")])

    # Create a treevie frame.
    tree_frame = Frame(show_suppliers_window)
    tree_frame.pack(fill="x", padx=20, pady=20)

    # Create a treeview Scrollbar
    tree_scroll = Scrollbar(tree_frame)
    tree_frame.pack(side=RIGHT, fill="y")

    # Create the columns.
    columns = ["1", "2"]

    # Create the treeview and pack it on the screen.
    treeview = ttk.Treeview(tree_frame, yscrollcommand=tree_scroll.set, selectmode="extended")
    treeview.pack(fill="x")

    # Configure the scrollbar.
    tree_scroll.config(command=treeview.yview)

    # Create striped row tags.
    treeview.tag_configure("evenrow", background="white")
    treeview.tag_configure("oddrow", background="lightblue")

    # Create the columns.
    columns = ["Network User CODE", "Network Supplier Name"]
    treeview.column("#0", width=0, stretch=NO)
    for i in range(0, len(columns)):
        treeview.column(i, anchor=W, width=140)

    # Create the headings.
    treeview.heading("#0", text="", anchor=W)
    for i in range(0, len(columns)):
        treeview.heading(i, text=str(columns[i]), anchor=W)

    # Set a counter.
    for key, value in data.items():
        treeview.insert(parent="", index="end", text="", values=(key, value))






# Create a tkinter main window.
root = Tk()
# Set window minimum size
root.minsize(width=400, height=400)
# Set window max size
# root.maxsize(width=250, height=250)
# Set the actual size.
root.geometry("250x250")
# Set the title.
root.title("Ordin 16")
# Set an icon
root.iconbitmap('images\\GazVest.ico')
# Options for month picking.
options = [
    "Ianuarie",
    "Februarie",
    "Martie",
    "Aprilie",
    "Mai",
    "Iunie",
    "Iulie",
    "August",
    "Septembrie",
    "Octombrie",
    "Noiembrie",
    "Decembrie"
]
# Configure a grid.
root.grid_rowconfigure(7, weight=1)
root.grid_columnconfigure(3, weight=1)

# Create a frame.
data_frame = LabelFrame(root, text="Import, Export, Generate")
data_frame.pack(fill="x", padx=2, pady=10)

database_frame = LabelFrame(root, text="Database Actions")
database_frame.pack(fill="x", padx=2, pady=10)

# Set a "Select the month" label for the user.
label = Label(data_frame, text="Select the month").grid(row=0, column=1, pady=10)
# Create a combobox with all months of an year.
combo = ttk.Combobox(data_frame, state="readonly", values=options)
combo.bind("<<ComboboxSelected>>", month_selected)
combo.grid(row=1, column=1, pady=2)
# Create a "Upload file" button to give to user the option to import a file.
import_btn = Button(data_frame, text="Import file", command=open_file).grid(row=2, column=0, sticky=W, padx=10, pady=10)
# Creat a button that asks the user where to save the files.
export_button = Button(data_frame, text="Save files to...", command=save_to).grid(row=2, column=2, sticky=E, padx=10, pady=10)
# Create a button that actually generates the needed files.
generate_button = Button(data_frame, text="Generate files", bg="green", command=generate_files).grid(row=3, column=1, pady=10)
# Add new supplier to all_suppliers.
new_sup_button = Button(database_frame, text="Add new supplier", fg="white", bg="blue", command=new_supplier).grid(row=4, column=0, sticky=W, pady=50)
# Add a delete supplier button.
delete_supplier = Button(database_frame, text="Delete Supplier", fg="white", bg="red", command=delete_supplier).grid(row=4, column=2, sticky=E, pady=2)
# Add a show all suppliers button.
show_suppliers = Button(database_frame, text="Show All Suppliers", bg="yellow", command=show_all_suppliers).grid(row=5, column=1, pady=2)

root.mainloop()



